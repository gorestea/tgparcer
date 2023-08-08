import json
import openpyxl

def write_excel():
    with open('result.json', 'r', encoding='utf-8') as f:
        result_dict = json.load(f)
    channels = result_dict.keys()
    for channel in channels:
        for word, messages in result_dict.items():
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = word
            sheet.column_dimensions['A'].width = 50
            sheet.column_dimensions['B'].width = 100
            sheet['A1'] = 'URL'
            sheet['B1'] = 'Text'
            row_count = 2
            for message_id, message_data in messages.items():
                sheet.cell(row=row_count, column=1).value = message_data['url']
                sheet.cell(row=row_count, column=2).value = message_data['text']
                sheet.cell(row=row_count, column=1).hyperlink = message_data['url']
                row_count += 1
            excel_filename = f"{word}.xlsx"
            wb.save(excel_filename)

if __name__ == '__main__':
    write_excel()



